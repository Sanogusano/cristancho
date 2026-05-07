import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from utils.triangulate import STATUS_LABELS, STATUS_COLORS


def _header_style(cell, bg="2C3E50", fg="FFFFFF"):
    cell.font = Font(bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_row_color(ws, row_idx: int, n_cols: int, hex_color: str):
    clean = hex_color.lstrip("#")
    fill = PatternFill("solid", start_color=clean)
    for col in range(1, n_cols + 1):
        ws.cell(row_idx, col).fill = fill


def export_triangulation(results: pd.DataFrame, meta: dict) -> bytes:
    """Genera Excel con resumen + tabla de todas las discrepancias."""
    wb = Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Resumen"

    ws_sum["A1"] = "Triangulación de Inventario — CEDI Guayabal"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A3"] = f'NetSuite: {meta.get("ns_filename", "")}'
    ws_sum["A4"] = f'Shopify:  {meta.get("sho_filename", "")}'
    ws_sum["A5"] = f'Fecha:    {meta.get("fecha", datetime.now().strftime("%Y-%m-%d %H:%M"))}'

    ws_sum["A7"] = "Estado"
    ws_sum["B7"] = "SKUs"
    _header_style(ws_sum["A7"])
    _header_style(ws_sum["B7"])

    status_order = ["ok", "ns_mayor", "sho_mayor", "sin_match_ns", "sin_match_sho"]
    counts = results["status"].value_counts().to_dict()

    for i, status in enumerate(status_order, 8):
        label = STATUS_LABELS.get(status, status)
        count = counts.get(status, 0)
        ws_sum[f"A{i}"] = label
        ws_sum[f"B{i}"] = count
        color = STATUS_COLORS.get(status, "#FFFFFF").lstrip("#")
        fill = PatternFill("solid", start_color=color)
        ws_sum[f"A{i}"].fill = fill
        ws_sum[f"B{i}"].fill = fill

    ws_sum.column_dimensions["A"].width = 38
    ws_sum.column_dimensions["B"].width = 10

    # ── Hoja 2: Todas las discrepancias ──────────────────────────────────────
    ws_disc = wb.create_sheet("Discrepancias")
    disc_headers = ["SKU", "Nombre", "Subtipo", "Línea", "Color", "Talla",
                    "NS Guayabal", "Shopify CEDI", "Diferencia", "Estado"]
    for col, h in enumerate(disc_headers, 1):
        _header_style(ws_disc.cell(1, col, h))
    ws_disc.row_dimensions[1].height = 30

    disc_data = results[results["status"] != "ok"].sort_values(
        "status", key=lambda s: s.map({"sho_mayor": 0, "ns_mayor": 1, "sin_match_ns": 2, "sin_match_sho": 3})
    )

    for r_idx, (_, row) in enumerate(disc_data.iterrows(), 2):
        ws_disc.cell(r_idx, 1, str(row["sku"]))
        ws_disc.cell(r_idx, 2, str(row.get("nombre") or ""))
        ws_disc.cell(r_idx, 3, str(row.get("subtipo") or ""))
        ws_disc.cell(r_idx, 4, str(row.get("linea") or ""))
        ws_disc.cell(r_idx, 5, str(row.get("color") or ""))
        ws_disc.cell(r_idx, 6, str(row.get("talla") or ""))
        ws_disc.cell(r_idx, 7, int(row["ns_guayabal"]))
        ws_disc.cell(r_idx, 8, int(row["sho_cedi"]))
        ws_disc.cell(r_idx, 9, int(row["diff"]))
        ws_disc.cell(r_idx, 10, STATUS_LABELS.get(row["status"], row["status"]))
        _apply_row_color(ws_disc, r_idx, 10, STATUS_COLORS.get(row["status"], "#FFFFFF"))

    for i, w in enumerate([20, 42, 14, 16, 18, 8, 14, 14, 12, 32], 1):
        ws_disc.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_resurtido(resurtido: pd.DataFrame, meta: dict) -> bytes:
    """Genera Excel de orden de resurtido PRINCIPAL / DISTRIBUIDORES → CEDI."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Orden de Resurtido"

    ws["A1"] = "ORDEN DE RESURTIDO — CEDI Guayabal"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        f'Fecha: {meta.get("fecha", datetime.now().strftime("%Y-%m-%d"))} | '
        "Origen: Bodega Principal / Distribuidores → Destino: CEDI Guayabal"
    )
    ws["A2"].font = Font(italic=True, color="555555", size=9)

    headers = [
        "SKU", "Nombre producto", "Subtipo",
        "Qty CEDI actual", "Qty Principal disp.", "Qty Distribuidor disp.",
        "Qty sugerida traslado", "Fuente",
    ]
    for col, h in enumerate(headers, 1):
        _header_style(ws.cell(4, col, h), bg="1A5276")
    ws.row_dimensions[4].height = 35

    for r_idx, (_, row) in enumerate(resurtido.sort_values("qty_sugerida", ascending=False).iterrows(), 5):
        ws.cell(r_idx, 1, str(row["sku"]))
        ws.cell(r_idx, 2, str(row.get("nombre") or ""))
        ws.cell(r_idx, 3, str(row.get("subtipo") or ""))
        ws.cell(r_idx, 4, int(row["ns_guayabal"]))
        ws.cell(r_idx, 5, int(row["ns_principal"]))
        ws.cell(r_idx, 6, int(row["ns_distribuidores"]))
        ws.cell(r_idx, 7, int(row["qty_sugerida"]))
        ws.cell(r_idx, 8, str(row["fuente"]))
        color = "D5F5E3" if row["fuente"] == "PRINCIPAL" else "D6EAF8"
        _apply_row_color(ws, r_idx, 8, f"#{color}")

    for i, w in enumerate([20, 45, 14, 14, 17, 19, 18, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
